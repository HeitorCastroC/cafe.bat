# -*- coding: utf-8 -*-
import os
import sys
from datetime import datetime
from pathlib import Path

# ── Configuração da planilha ───────────────────────────────────────────────────
PASTA      = Path(__file__).parent
PLANILHA   = PASTA / "escala_cafe.xlsx"

# ── Carregamento dinâmico dos dados ───────────────────────────────────────────
def carregar_escala():
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("\n  ✗  Biblioteca 'openpyxl' não encontrada.")
        print("     Instale com:  pip install openpyxl\n")
        sys.exit(1)

    if not PLANILHA.exists():
        print(f"\n  ✗  Planilha não encontrada: {PLANILHA}")
        print("     Certifique-se de que 'escala_cafe.xlsx' está na mesma pasta do script.\n")
        sys.exit(1)

    wb = load_workbook(PLANILHA, data_only=True)
    ws = wb.active
    escala = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome, data, pacotes, concluido = (list(row) + [None, None, None, None])[:4]
        if not nome or not data:
            continue
        # Normaliza data: aceita objeto date/datetime ou string
        if hasattr(data, "strftime"):
            data = data.strftime("%d/%m/%Y")
        else:
            data = str(data).strip()
        pacotes   = int(pacotes)   if pacotes   is not None else 0
        concluido = bool(concluido) if concluido is not None else False
        escala.append((str(nome).strip(), data, pacotes, concluido))
    return escala

# ── Instruções ────────────────────────────────────────────────────────────────
INSTRUCOES = [
    "1  Desligue a garrafa de café",
    "2  Retire o pó usado e o café que sobrou do dia anterior",
    "3  Lave o coador",
    "4  Complete com água filtrada até a marca MAX (6L) da garrafa",
    "5  Insira o pino e o coador novamente na garrafa",
    "6  Coloque 16 colheres cheias de pó de café no coador (para 6L)",
    "7  Feche e ligue a cafeteira",
    "8  Informe o horário que ficará pronto (~30 min)",
]

MARCAS       = "3 Corações (Forte, Extra Forte, Tradicional)  •  Melita"
INGREDIENTES = "Pó de café · 16 medidas · Água 6L (MAX) · ~30 min"

# ── Helpers ────────────────────────────────────────────────────────────────────
DIAS_PT  = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
MESES_PT = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]

def limpar():
    os.system("cls" if os.name == "nt" else "clear")

def linha(char="─", n=68):
    return char * n

def centralizar(texto, n=68):
    return texto.center(n)

def parse_data(s):
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def formatar_data(dt):
    return f"{dt.day:02d}/{MESES_PT[dt.month-1]}/{dt.year}  ({DIAS_PT[dt.weekday()]})"

def colheres_label(n):
    if n == 0:
        return "0 pacotes doados :("
    elif n == 1:
        return "1 pacote doado"
    else:
        return f"{n} pacotes doados"

# ── ASCII Art ──────────────────────────────────────────────────────────────────
HEADER = r'''
                        (
                          )     (
                   ___...(-------)-....___
               .-""       )    (          ""-.
         .-'``'|-._             )         _.-|
        /  .--.|   `""---...........---""`   |
       /  /    |                             |
       |  |    |                             |
        \  \   |           Bom dia!          |
         `\ `\ |             FSFX            |
           `\ `|                             |
           _/ /\                             /
          (__/  \                           /
       _..---""` \                         /`""---.._
    .-'           \                       /          '-.
   :               `-.__             __.-'              :
   :                  ) ""---...---"" (                 :
    '._               `"--...___...--"`              _.'
      \""--..__                              __..--""/
       '._     """----.....______.....----"""     _.'
          `""--..,,_____            _____,,..--""`
                        `"""----"""`
'''

TITULO = [
    "╔══════════════════════════════════════════════════════════════════╗",
    "║      ☕  ESCALA DE CAFÉ  —  FUNDAÇÃO SÃO FRANCISCO XAVIER  ☕   ║",
    "║                    REV. 1  ·  03/26 - 06/26                      ║",
    "╚══════════════════════════════════════════════════════════════════╝",
]

def imprimir_ascii(texto):
    linhas = texto.splitlines()
    largura_max = max((len(l) for l in linhas if l.strip()), default=0)
    offset = max((68 - largura_max) // 2, 0)
    for l in linhas:
        print(" " * offset + l)

def imprimir_cabecalho(n_registros):
    limpar()
    imprimir_ascii(HEADER)
    for t in TITULO:
        print(centralizar(t))
    print(centralizar(f"[ planilha: escala_cafe.xlsx  ·  {n_registros} registros ]"))
    print()

def imprimir_instrucoes():
    print(centralizar("┌─────────────────────────────────────┐"))
    print(centralizar("│   PROCEDIMENTO PARA FAZER O CAFÉ    │"))
    print(centralizar("└─────────────────────────────────────┘"))
    print()
    for passo in INSTRUCOES:
        print(f"    {passo}")
    print()
    print(f"  {'Marcas    :'} {MARCAS}")
    print(f"  {'Info      :'} {INGREDIENTES}")
    print()
    print(linha())

# ── Busca ──────────────────────────────────────────────────────────────────────
def buscar_por_nome(escala, termo):
    termo = termo.strip().lower()
    resultados = [(n, d, c) for n, d, c, _ in escala if termo in n.lower()]
    if not resultados:
        print(f"\n  ✗  Nenhum resultado para \"{termo}\".\n")
        return
    print(f"\n  ✔  {len(resultados)} entrada(s) encontrada(s):\n")
    print(f"  {'NOME':<22} {'DATA':<30} {'PACOTES'}")
    print(f"  {linha('-', 60)}")
    for nome, data_str, col in resultados:
        dt = parse_data(data_str)
        data_fmt = formatar_data(dt) if dt else data_str
        print(f"  {nome:<22} {data_fmt:<30} {colheres_label(col)}")
    print()

def buscar_por_data(escala, entrada):
    entrada = entrada.strip()
    dt_busca = parse_data(entrada)
    if not dt_busca:
        try:
            dt_busca = datetime.strptime(entrada + "/2026", "%d/%m/%Y")
        except ValueError:
            print("\n  ✗  Data inválida. Use o formato DD/MM ou DD/MM/AAAA.\n")
            return

    dia_semana = dt_busca.weekday()
    if dia_semana >= 5:
        nome_dia = "sábado" if dia_semana == 5 else "domingo"
        print(f"\n  ☕  Hoje é {nome_dia.upper()}... café só na casa de mamãe! 🏠\n")
        return

    data_fmt = formatar_data(dt_busca)
    encontrados = []
    for nome, data_str, col, _ in escala:
        dt = parse_data(data_str)
        if dt and dt.date() == dt_busca.date():
            encontrados.append((nome, col))

    if not encontrados:
        print(f"\n  ✗  Nenhum responsável cadastrado para {data_fmt}.\n")
        return

    print(f"\n  📅  {data_fmt}\n")
    print(f"  {'RESPONSÁVEL':<25} {'PACOTES'}")
    print(f"  {linha('-', 45)}")
    for nome, col in encontrados:
        print(f"  {nome:<25} {colheres_label(col)}")
    print()

# ── Loop principal ─────────────────────────────────────────────────────────────
def main():
    while True:
        escala = carregar_escala()          # relê a planilha a cada iteração
        imprimir_cabecalho(len(escala))
        imprimir_instrucoes()
        print()
        print("  PESQUISA")
        print(f"  {linha('-', 40)}")
        print("  [1]  Buscar por NOME")
        print("  [2]  Buscar por DATA  (DD/MM ou DD/MM/AAAA)")
        print("  [0]  Sair")
        print()

        opcao = input("  Opção: ").strip()
        print()

        if opcao == "1":
            termo = input("  Digite o nome (parcial): ").strip()
            buscar_por_nome(escala, termo)

        elif opcao == "2":
            entrada = input("  Digite a data (ex: 19/06 ou 19/06/2026): ").strip()
            buscar_por_data(escala, entrada)

        elif opcao == "0":
            limpar()
            print("\n  Até logo! ☕\n")
            sys.exit(0)

        else:
            print("  ✗  Opção inválida.\n")

        input("  Pressione ENTER para voltar ao menu...")

if __name__ == "__main__":
    main()